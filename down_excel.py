#!/usr/bin/env python
# -*- coding: utf-8 -*-
import imghdr
import os
import eventlet
import random
import requests
from documents import connect, jumia
from openpyxl.drawing.image import Image
from openpyxl import Workbook
connect()
eventlet.monkey_patch(socket=True, select=True)
requests = eventlet.import_patched('requests.__init__')
connect()
import time


def start_excel(key):
    a = jumia.objects(key=key).order_by("-comment")[0:300]
    print(a)
    photo = []
    A=1
    wb = Workbook()
    for i in a:
        while True:
            try:
                response = requests.get(i.photo)
                img = response.content
                with open('./img/{0}.jpg'.format(i.photo.split('/')[3]), 'wb') as f:
                    f.write(img)
                break
            except:
                break

        print(i.comment, i.price, i.photo)
        ws = wb.active
        ws.column_dimensions['A'].width = 45  # 修改列A的列宽
        ws.row_dimensions[A].height = 45  # 修改行3的行高

        price = i.price
        ws['D{0}'.format(A)] = price
        ws['A{0}'.format(A)] = str(i.id)
        ws['B{0}'.format(A)] = i.name
        ws['C{0}'.format(A)] = i.comment

        ws['G{0}'.format(A)] = i.url
        # Rows can also be appended
        # ws.append([1, 2, 3])
        # Python types will automatically be converted
        if i.photo=='XXXX':
            import requests
            # requests.get()
            pass
        else:
            try:
                if imghdr.what('./img/{0}.jpg'.format(i.photo.split('/')[3])) == 'jpeg':
                    img = Image('./img/{0}.jpg'.format(i.photo.split('/')[3]))
                    img.width, img.height = (60, 60)
                    # 这两个属性分别是对应添加图片的宽高    
                    ws.add_image(img, 'F{0}'.format(A))
            except:
                pass
            # Save the file
        A+=1
    wb.save("./log/{0}.xlsx".format(key+str(time.time())))
start_excel('ci')
